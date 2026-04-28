from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.protection import WorkbookProtection

from app.models.order import ERPRow, Order, OrderItem
from app.utils.logger import logger

HEADERS = [
    "PEDIDO",
    "NOME_CLIENTE",
    "CNPJ_CLIENTE",
    "CODIGO_PRODUTO",
    "EAN",
    "DESCRICAO",
    "QUANTIDADE",
    "PRECO_UNITARIO",
    "VALOR_TOTAL",
    "OBS",
    "DATA_ENTREGA",
    "CNPJ_LOCAL_ENTREGA",
    "EAN_LOCAL_ENTREGA",
]

HEADER_COLOR = "1F4E79"
COL_WIDTHS = [18, 40, 20, 18, 16, 50, 12, 16, 16, 35, 16, 22, 18]


class ERPExporter:
    def export(self, order: Order, output_dir: str = "output") -> list[Path]:
        """Export order. Splits into multiple files if delivery locations differ."""
        groups = self._group_by_delivery(order)

        if len(groups) <= 1:
            items = list(groups.values())[0] if groups else order.items
            return [self._write_file(order, items, output_dir, suffix=None)]

        paths = []
        for i, (key, items) in enumerate(sorted(groups.items()), 1):
            suffix = self._suffix_for_group(items, fallback=str(i))
            paths.append(self._write_file(order, items, output_dir, suffix=suffix))
        return paths

    def _suffix_for_group(self, items: list[OrderItem], fallback: str) -> str:
        """Sufixo legível para o nome do arquivo de cada loja."""
        first = items[0] if items else None
        if not first:
            return fallback

        # Sam's Club: split por EAN do local de entrega → "SAMS_LOJA_<filial-cnpj>"
        if first.delivery_ean:
            cnpj_digits = re.sub(r"[^\d]", "", first.delivery_cnpj or "")
            if len(cnpj_digits) >= 6:
                # últimos 6 dígitos = "0094" + "08" → "0094_08"
                tail = f"{cnpj_digits[-6:-2]}_{cnpj_digits[-2:]}"
                return f"SAMS_LOJA_{tail}"
            # sem CNPJ: usar últimos 4 do EAN
            return f"SAMS_LOJA_{first.delivery_ean[-4:]}"

        # Lojas só com nome (NBA): sanitizar delivery_name
        if first.delivery_name and not first.delivery_cnpj:
            return re.sub(r"[^\w\s]", "_", first.delivery_name).strip("_")[:30]

        return fallback

    # ------------------------------------------------------------------
    # Grouping
    # ------------------------------------------------------------------

    def _group_by_delivery(self, order: Order) -> dict[str, list[OrderItem]]:
        """Group items by distinct delivery CNPJ.
        Items with no delivery CNPJ or same as customer CNPJ go into a single group (key='').
        Only split when multiple distinct delivery locations exist.
        """
        buckets: dict[str, list[OrderItem]] = {}
        for item in order.items:
            key = self._delivery_key(item, order.header.customer_cnpj)
            buckets.setdefault(key, []).append(item)

        # If all items share the same key → no split needed
        if len(buckets) == 1:
            return buckets

        # Multiple locations: only keep the split if at least one has a real CNPJ
        real_keys = [k for k in buckets if k]
        if not real_keys:
            return {"": order.items}

        return buckets

    def _delivery_key(self, item: OrderItem, customer_cnpj: Optional[str]) -> str:
        """Chave de agrupamento por destino.

        Prioridade:
        1. `delivery_ean` (Sam's GRADE) — cada loja tem EAN único, evita
           ambiguidade quando a filial coincide com o CNPJ do customer.
        2. `delivery_cnpj` distinto do customer (Riachuelo).
        3. `delivery_name` (NBA, lojas identificadas só por nome).
        4. vazio (sem split).
        """
        if item.delivery_ean:
            return f"ean:{item.delivery_ean}"
        d = item.delivery_cnpj
        if d:
            d_digits = re.sub(r"[^\d]", "", d)
            c_digits = re.sub(r"[^\d]", "", customer_cnpj or "")
            if d_digits and d_digits == c_digits:
                return ""
            return d
        return item.delivery_name or ""

    # ------------------------------------------------------------------
    # File writing
    # ------------------------------------------------------------------

    def _write_file(
        self,
        order: Order,
        items: list[OrderItem],
        output_dir: str,
        suffix: Optional[str],
    ) -> Path:
        rows = self._to_erp_rows(order, items)
        filename = self._make_filename(order, suffix, items)
        path = Path(output_dir) / filename
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pedido"

        self._write_header(ws)
        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row.pedido)
            ws.cell(row=row_idx, column=2, value=row.nome_cliente)
            ws.cell(row=row_idx, column=3, value=row.cnpj_cliente)
            ws.cell(row=row_idx, column=4, value=row.codigo_produto)
            ws.cell(row=row_idx, column=5, value=row.ean)
            ws.cell(row=row_idx, column=6, value=row.descricao)
            ws.cell(row=row_idx, column=7, value=row.quantidade)
            ws.cell(row=row_idx, column=8, value=row.preco_unitario)
            ws.cell(row=row_idx, column=9, value=row.valor_total)
            ws.cell(row=row_idx, column=10, value=row.obs)
            ws.cell(row=row_idx, column=11, value=row.data_entrega)
            ws.cell(row=row_idx, column=12, value=row.cnpj_local_entrega)
            ws.cell(row=row_idx, column=13, value=row.ean_local_entrega)

        self._unlock_workbook(wb)
        wb.save(path)
        logger.info(f"Exportado: {path.name} ({len(rows)} linha(s))")
        return path

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _unlock_workbook(wb) -> None:
        """Garante que o XLSX gerado fique totalmente livre: sem senha, sem
        lockStructure/lockWindows/lockRevision, sem proteção de sheet e sem
        read-only-recommended (fileSharing). Defensivo contra regressões."""
        wb.security = WorkbookProtection()
        for ws in wb.worksheets:
            ws.protection.disable()
            ws.protection.sheet = False
            ws.protection.enabled = False
            ws.protection.hashValue = None
            ws.protection.saltValue = None
            ws.protection.algorithmName = None
            ws.protection.spinCount = None

    def _write_header(self, ws) -> None:
        fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = w

    def _make_filename(
        self, order: Order, suffix: Optional[str], items: Optional[list] = None
    ) -> str:
        name = order.header.customer_name or "SEM_CLIENTE"
        cnpj = re.sub(r"[^\d]", "", order.header.customer_cnpj or "")
        # When no header CNPJ (Riachuelo), use first item's delivery CNPJ
        if not cnpj and items:
            first_d = next((i.delivery_cnpj for i in items if i.delivery_cnpj), None)
            cnpj = re.sub(r"[^\d]", "", first_d or "")
        num = order.header.order_number or "SEM_NUMERO"
        name_clean = re.sub(r"[^\w]", "_", name).strip("_")[:30]
        cnpj_part = f"_{cnpj}" if cnpj else ""
        suffix_part = f"_{suffix}" if suffix else ""
        return f"{name_clean}{cnpj_part}_Pedido_{num}{suffix_part}.xlsx"

    def _to_erp_rows(self, order: Order, items: list[OrderItem]) -> list[ERPRow]:
        customer_cnpj_digits = re.sub(r"[^\d]", "", order.header.customer_cnpj or "")
        # When order has no header-level CNPJ (e.g. Riachuelo), use delivery CNPJ as customer
        use_delivery_as_customer = not order.header.customer_cnpj

        rows = []
        for item in items:
            d = item.delivery_cnpj or ""
            d_digits = re.sub(r"[^\d]", "", d)

            if use_delivery_as_customer and d:
                # Riachuelo: loja = cliente (has delivery CNPJ)
                effective_customer_name = item.delivery_name or order.header.customer_name
                effective_customer_cnpj = d
                delivery_cnpj_col = d  # repeat in local_entrega too
            elif use_delivery_as_customer and item.delivery_name and not d:
                # NBA-style: loja identificada só por nome, sem CNPJ
                effective_customer_name = item.delivery_name
                effective_customer_cnpj = None
                delivery_cnpj_col = None
            else:
                effective_customer_name = order.header.customer_name
                effective_customer_cnpj = order.header.customer_cnpj
                # Leave blank if same as customer or missing
                delivery_cnpj_col = d if (d_digits and d_digits != customer_cnpj_digits) else None

            rows.append(
                ERPRow(
                    pedido=order.header.order_number or "SEM_NUMERO",
                    nome_cliente=effective_customer_name,
                    cnpj_cliente=effective_customer_cnpj,
                    codigo_produto=item.product_code,
                    ean=item.ean,
                    descricao=item.description or "",
                    quantidade=item.quantity or 0.0,
                    preco_unitario=item.unit_price,
                    valor_total=item.total_price,
                    obs=item.obs,
                    data_entrega=item.delivery_date,
                    cnpj_local_entrega=delivery_cnpj_col,
                    ean_local_entrega=item.delivery_ean,
                )
            )
        return rows
