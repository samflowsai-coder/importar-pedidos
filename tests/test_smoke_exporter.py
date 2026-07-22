"""Smoke tests for app.exporters.erp_exporter.

Validates the public contract: filename composition, multi-store split,
and ERPRow generation. Pure unit tests — no I/O dependencies beyond a tmp dir.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from app.exporters.erp_exporter import ERPExporter
from app.models.order import Order, OrderHeader, OrderItem


def _order(header_kwargs=None, items=None) -> Order:
    return Order(
        header=OrderHeader(**(header_kwargs or {})),
        items=items or [],
    )


def test_export_single_store_writes_one_file(tmp_path: Path) -> None:
    order = _order(
        header_kwargs={
            "order_number": "PED-001",
            "customer_name": "CLIENTE TESTE",
            "customer_cnpj": "12.345.678/0001-90",
        },
        items=[
            OrderItem(description="TENIS A", quantity=10, unit_price=99.9),
            OrderItem(description="TENIS B", quantity=5, unit_price=149.9),
        ],
    )

    paths = ERPExporter().export(order, output_dir=str(tmp_path))

    assert len(paths) == 1
    assert paths[0].exists()
    assert paths[0].suffix == ".xlsx"
    # Filename pattern: {NOME}_{CNPJ}_Pedido_{NUM}.xlsx
    assert "CLIENTE_TESTE" in paths[0].name
    assert "12345678000190" in paths[0].name
    assert "PED-001" in paths[0].name


def test_export_splits_by_distinct_delivery_cnpj(tmp_path: Path) -> None:
    order = _order(
        header_kwargs={
            "order_number": "PED-002",
            "customer_name": "RIACHUELO",
            "customer_cnpj": None,  # Riachuelo: no header CNPJ → delivery is the customer
        },
        items=[
            OrderItem(description="A", quantity=1, delivery_cnpj="11.111.111/0001-11"),
            OrderItem(description="B", quantity=2, delivery_cnpj="22.222.222/0001-22"),
        ],
    )

    paths = ERPExporter().export(order, output_dir=str(tmp_path))

    assert len(paths) == 2, "two distinct delivery CNPJs → two files"
    for p in paths:
        assert p.exists()


def test_export_groups_same_cnpj_into_single_file(tmp_path: Path) -> None:
    cnpj = "99.999.999/0001-99"
    order = _order(
        header_kwargs={"order_number": "PED-003", "customer_name": "X", "customer_cnpj": cnpj},
        items=[
            OrderItem(description="A", quantity=1, delivery_cnpj=cnpj),
            OrderItem(description="B", quantity=2, delivery_cnpj=cnpj),
        ],
    )

    paths = ERPExporter().export(order, output_dir=str(tmp_path))
    assert len(paths) == 1, "delivery==customer → no split"


def test_export_splits_by_delivery_name_when_no_cnpj(tmp_path: Path) -> None:
    """NBA-style: stores identified by name only."""
    order = _order(
        header_kwargs={"order_number": "PED-004", "customer_name": "NBA", "customer_cnpj": None},
        items=[
            OrderItem(description="A", quantity=1, delivery_name="LOJA SP"),
            OrderItem(description="B", quantity=2, delivery_name="LOJA RJ"),
        ],
    )

    paths = ERPExporter().export(order, output_dir=str(tmp_path))
    assert len(paths) == 2
    suffixes = sorted(p.stem.split("_")[-1] for p in paths)
    # Suffix derived from delivery_name (sanitized)
    assert any("LOJA" in s or s in {"SP", "RJ"} for s in suffixes)


def test_export_strips_illegal_control_chars(tmp_path: Path) -> None:
    """Um caractere de controle no texto do pedido (vindo de parse de XLS/PDF
    sujo) NÃO pode derrubar a geração do XLS. openpyxl recusa 0x00–0x1F e
    levantaria IllegalCharacterError → HTTP 500 na rota /export-xlsx.

    Regressão: pedido AF185/H2S4 (2026-07-22) — 1 item com char invisível.
    """
    order = _order(
        header_kwargs={"order_number": "AF185", "customer_name": "H2S4", "customer_cnpj": "111"},
        items=[
            OrderItem(
                description="KIT 3 PARES - SAPATILHA - BRANCO \x1f- 33 - 38",
                quantity=12,
                unit_price=11.96,
                obs="obs\x00suja",
            ),
        ],
    )

    # Não pode levantar exceção.
    [path] = ERPExporter().export(order, output_dir=str(tmp_path))

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # Char ilegal removido; o resto do texto permanece legível.
    assert ws.cell(row=2, column=6).value == "KIT 3 PARES - SAPATILHA - BRANCO - 33 - 38"
    assert ws.cell(row=2, column=10).value == "obssuja"


def test_export_preserves_legit_whitespace(tmp_path: Path) -> None:
    """Sanitização remove só controle ilegal — tab/newline/CR são válidos no XLSX
    e devem sobreviver."""
    order = _order(
        header_kwargs={"order_number": "PED-006", "customer_name": "X", "customer_cnpj": "111"},
        items=[OrderItem(description="LINHA1\nLINHA2\tFIM", quantity=1)],
    )
    [path] = ERPExporter().export(order, output_dir=str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert wb.active.cell(row=2, column=6).value == "LINHA1\nLINHA2\tFIM"


def test_export_produces_readable_xlsx_with_headers(tmp_path: Path) -> None:
    order = _order(
        header_kwargs={"order_number": "PED-005", "customer_name": "X", "customer_cnpj": "111"},
        items=[OrderItem(description="ITEM", quantity=3, unit_price=10.0, total_price=30.0)],
    )
    [path] = ERPExporter().export(order, output_dir=str(tmp_path))
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, 13)]
    assert headers[0] == "PEDIDO"
    assert headers[5] == "DESCRICAO"
    assert ws.cell(row=2, column=6).value == "ITEM"
    assert ws.cell(row=2, column=7).value == 3
