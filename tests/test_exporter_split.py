"""Cobertura do split por loja e da coluna EAN_LOCAL_ENTREGA no exportador."""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from app.exporters.erp_exporter import HEADERS

SAMPLES = Path(__file__).parent.parent / "samples"


def _load(filename: str):
    from app.ingestion.file_loader import LoadedFile

    p = SAMPLES / filename
    if not p.exists():
        pytest.skip(f"Sample not found: {filename}")
    return LoadedFile(path=p, extension=p.suffix.lower(), raw=p.read_bytes())


def _process(filename: str):
    from app.pipeline import process

    return process(_load(filename))


def _open_xlsx(path: Path):
    wb = openpyxl.load_workbook(path)
    return wb.active


def test_headers_include_ean_local_entrega():
    """O header do XLSX deve incluir EAN_LOCAL_ENTREGA na 13ª coluna."""
    assert "EAN_LOCAL_ENTREGA" in HEADERS
    assert HEADERS.index("EAN_LOCAL_ENTREGA") == 12  # 0-indexed → 13ª coluna


def test_sams_grade_export_creates_three_files():
    from app.exporters.erp_exporter import ERPExporter

    order = _process("PEDIDO SAMS CLUB GRADE.pdf")
    with tempfile.TemporaryDirectory() as tmp:
        paths = ERPExporter().export(order, tmp)
        assert len(paths) == 3


def test_sams_grade_export_xlsx_columns():
    """XLSX exportado tem todos os 13 cabeçalhos esperados."""
    from app.exporters.erp_exporter import ERPExporter

    order = _process("PEDIDO SAMS CLUB GRADE.pdf")
    with tempfile.TemporaryDirectory() as tmp:
        paths = ERPExporter().export(order, tmp)
        ws = _open_xlsx(paths[0])
        headers = [c.value for c in ws[1]]
        assert headers == HEADERS


def test_sams_grade_ean_local_populated_in_xlsx():
    """Coluna EAN_LOCAL_ENTREGA preenchida com EAN da loja em todas as linhas."""
    from app.exporters.erp_exporter import ERPExporter

    order = _process("PEDIDO SAMS CLUB GRADE.pdf")
    with tempfile.TemporaryDirectory() as tmp:
        paths = ERPExporter().export(order, tmp)
        for path in paths:
            ws = _open_xlsx(path)
            ean_col = HEADERS.index("EAN_LOCAL_ENTREGA") + 1  # 1-indexed
            file_eans = {ws.cell(row=r, column=ean_col).value for r in range(2, ws.max_row + 1)}
            # Cada arquivo tem exatamente 1 EAN de loja
            assert len(file_eans) == 1
            (ean,) = file_eans
            assert ean is not None and len(ean) == 13


def test_sams_grade_qty_per_file_matches_grade():
    """Soma de QUANTIDADE em cada arquivo bate com a soma daquela loja na grade."""
    from app.exporters.erp_exporter import ERPExporter

    order = _process("PEDIDO SAMS CLUB GRADE.pdf")
    expected = {
        "7891737001698": sum(i.quantity for i in order.items if i.delivery_ean == "7891737001698"),
        "7891737012779": sum(i.quantity for i in order.items if i.delivery_ean == "7891737012779"),
        "7891737676568": sum(i.quantity for i in order.items if i.delivery_ean == "7891737676568"),
    }
    with tempfile.TemporaryDirectory() as tmp:
        paths = ERPExporter().export(order, tmp)
        ean_col = HEADERS.index("EAN_LOCAL_ENTREGA") + 1
        qty_col = HEADERS.index("QUANTIDADE") + 1
        for path in paths:
            ws = _open_xlsx(path)
            file_ean = ws.cell(row=2, column=ean_col).value
            total = sum(ws.cell(row=r, column=qty_col).value or 0 for r in range(2, ws.max_row + 1))
            assert total == expected[file_ean]


def test_sams_consolidated_still_single_file():
    """Regressão: o sample consolidado (sem GRADE) continua gerando 1 só arquivo."""
    from app.exporters.erp_exporter import ERPExporter

    order = _process("PEDIDO SAMS CLUB.pdf")
    with tempfile.TemporaryDirectory() as tmp:
        paths = ERPExporter().export(order, tmp)
        assert len(paths) == 1


def test_export_has_no_workbook_protection():
    """XLSX gerado nunca deve ter senha, lockStructure, lockWindows ou lockRevision."""
    from app.exporters.erp_exporter import ERPExporter

    order = _process("PEDIDO SAMS CLUB GRADE.pdf")
    with tempfile.TemporaryDirectory() as tmp:
        paths = ERPExporter().export(order, tmp)
        for p in paths:
            wb = openpyxl.load_workbook(p)
            sec = wb.security
            assert not sec.workbookPassword
            assert not sec.lockStructure
            assert not sec.lockWindows
            assert not sec.lockRevision
            assert not sec.revisionsPassword
            assert wb.read_only is False


def test_export_has_no_sheet_protection():
    """Cada sheet do XLSX gerado deve estar destravada (sem senha, sem sheet=True)."""
    from app.exporters.erp_exporter import ERPExporter

    order = _process("PEDIDO SAMS CLUB GRADE.pdf")
    with tempfile.TemporaryDirectory() as tmp:
        paths = ERPExporter().export(order, tmp)
        for p in paths:
            wb = openpyxl.load_workbook(p)
            for ws in wb.worksheets:
                assert ws.protection.sheet is False
                assert ws.protection.enabled is False
                assert not ws.protection.hashValue
                assert not ws.protection.saltValue


def test_export_has_no_read_only_recommended():
    """Não deve haver fileSharing/readOnlyRecommended no XML do workbook."""
    import zipfile

    from app.exporters.erp_exporter import ERPExporter

    order = _process("PEDIDO SAMS CLUB GRADE.pdf")
    with tempfile.TemporaryDirectory() as tmp:
        paths = ERPExporter().export(order, tmp)
        for p in paths:
            with zipfile.ZipFile(p) as z:
                xml = z.read("xl/workbook.xml").decode("utf-8")
                assert "fileSharing" not in xml
                assert "readOnlyRecommended" not in xml
